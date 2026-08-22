"""
Genel Excel (.xlsx) üretici — Sadoksan python-service

Neden: CSV çıktıları Türkçe/Avrupa yerel ayarlı Windows Excel'de tek sütunda
(A sütununda virgüllerle birleşik) açılıyordu. 'sep=,' ipucu her sürümde
çalışmadığı gibi çevrimiçi editörlerde de fazladan satır olarak görünüyordu.
Gerçek .xlsx üretildiğinde ayraç sorunu tamamen ortadan kalkar; sütunlar her
yerde doğru ayrılır, sayı/tarih/para biçimleri korunur.

Kullanım (app.py içinden):
    from excel_generator import ExcelGenerator
    buf = ExcelGenerator().generate(sheets=[...], title="Ürünler")
"""

from io import BytesIO
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Sütun tipine göre Excel biçim kodları
FORMATS = {
    "text": "@",
    "number": "#,##0",
    "decimal": "#,##0.00",
    "money": '#,##0.00 "TL"',
    "usd": '"$"#,##0.00',
    "percent": "0.00%",
    "date": "DD.MM.YYYY",
    "datetime": "DD.MM.YYYY HH:MM",
}

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


class ExcelGenerator:
    """JSON tanımından biçimlendirilmiş .xlsx üretir."""

    MAX_WIDTH = 60
    MIN_WIDTH = 8

    def generate(self, sheets, title=None):
        """
        sheets: [{
          "name": "Ürünler",
          "columns": [{"key": "sku", "label": "SKU", "type": "text", "width": 18}, ...],
          "rows": [{"sku": "A123", ...}, ...]   veya  [[...], [...]]
        }]
        """
        wb = Workbook()
        wb.remove(wb.active)

        if not sheets:
            sheets = [{"name": "Sayfa1", "columns": [], "rows": []}]

        for sheet in sheets:
            self._build_sheet(wb, sheet)

        if title:
            wb.properties.title = title
        wb.properties.creator = "Sadoksan ERP"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ------------------------------------------------------------------
    def _build_sheet(self, wb, sheet):
        name = (sheet.get("name") or "Sayfa1")[:31]
        # Excel sayfa adında bu karakterler yasak
        for ch in "[]:*?/\\":
            name = name.replace(ch, "-")
        ws = wb.create_sheet(title=name)

        columns = sheet.get("columns") or []
        rows = sheet.get("rows") or []

        # Sütun tanımı verilmediyse ilk satırdan türet
        if not columns and rows and isinstance(rows[0], dict):
            columns = [{"key": k, "label": str(k)} for k in rows[0].keys()]

        # --- başlık satırı ---
        for i, col in enumerate(columns, start=1):
            c = ws.cell(row=1, column=i, value=col.get("label") or col.get("key"))
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER

        genislikler = [len(str(c.get("label") or "")) for c in columns]

        # --- veri satırları ---
        for r, row in enumerate(rows, start=2):
            for i, col in enumerate(columns, start=1):
                key = col.get("key")
                deger = row.get(key) if isinstance(row, dict) else (
                    row[i - 1] if i - 1 < len(row) else None
                )
                deger = self._normalize(deger, col.get("type"))
                cell = ws.cell(row=r, column=i, value=deger)
                cell.border = BORDER

                tip = col.get("type") or self._infer(deger)
                if tip in FORMATS:
                    cell.number_format = FORMATS[tip]
                if tip in ("number", "decimal", "money", "usd", "percent"):
                    cell.alignment = Alignment(horizontal="right")
                elif tip in ("date", "datetime"):
                    cell.alignment = Alignment(horizontal="center")

                uzunluk = len(str(deger)) if deger is not None else 0
                if i - 1 < len(genislikler):
                    genislikler[i - 1] = max(genislikler[i - 1], min(uzunluk, self.MAX_WIDTH))

        # --- sütun genişlikleri ---
        for i, col in enumerate(columns, start=1):
            w = col.get("width") or (genislikler[i - 1] + 3 if i - 1 < len(genislikler) else 14)
            ws.column_dimensions[get_column_letter(i)].width = max(self.MIN_WIDTH, min(w, self.MAX_WIDTH))

        # --- başlığı dondur + otomatik filtre ---
        if columns:
            ws.freeze_panes = "A2"
            son = f"{get_column_letter(len(columns))}{max(len(rows) + 1, 1)}"
            ws.auto_filter.ref = f"A1:{son}"
        ws.sheet_view.showGridLines = False

    # ------------------------------------------------------------------
    @staticmethod
    def _normalize(deger, tip=None):
        """Excel'in sayı/tarih olarak tanıması için değeri dönüştürür."""
        if deger is None or deger == "":
            return None
        if isinstance(deger, bool):
            return "Evet" if deger else "Hayır"
        if tip in ("number", "decimal", "money", "usd", "percent"):
            try:
                return float(deger)
            except (TypeError, ValueError):
                return deger
        if tip in ("date", "datetime") and isinstance(deger, str):
            metin = deger.replace("Z", "+00:00")
            try:
                return datetime.fromisoformat(metin).replace(tzinfo=None)
            except ValueError:
                return deger
        return deger

    @staticmethod
    def _infer(deger):
        if isinstance(deger, (datetime, date)):
            return "datetime" if isinstance(deger, datetime) else "date"
        if isinstance(deger, int):
            return "number"
        if isinstance(deger, float):
            return "decimal"
        return "text"
